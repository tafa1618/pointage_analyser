"""
Onglet Exhaustivité — Calendrier de présence technicien × jour.
Optimisé pour l'export Excel multi-équipes.
"""

from __future__ import annotations
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from io import BytesIO

from pointage_analyzer.utils.senegal_holidays import (
    nb_jours_ouvres as _nb_jours_ouvres,
    get_holidays_range as _get_feries,
)

from pointage_analyzer.pipeline.exhaustivite_builder import (
    ExhaustiviteBuilder,
    ExhaustiviteError,
    PresenceStatus,
)
from pointage_analyzer.core.config import ScoringConfig

# Palette couleurs sémantiques
STATUS_COLORS = {
    PresenceStatus.PRESENT:      "#2ECC71",
    PresenceStatus.ABSENT:       "#E74C3C",
    PresenceStatus.EXCESSIF:     "#E67E22",
    PresenceStatus.WEEKEND:      "#BDC3C7",
    PresenceStatus.FERIE:        "#85C1E9",
    PresenceStatus.NON_CONCERNE: "#FFFFFF",
}

STATUS_LABELS = {
    PresenceStatus.PRESENT:   "Présent (≤ 8h)",
    PresenceStatus.ABSENT:    "Absent (0h)",
    PresenceStatus.EXCESSIF:  "Excessif (> 8h)",
    PresenceStatus.WEEKEND:   "Week-end",
    PresenceStatus.FERIE:     "Jour férié",
}

_JOURS_FR = {0: "Lun", 1: "Mar", 2: "Mer", 3: "Jeu", 4: "Ven", 5: "Sam", 6: "Dim"}


def render_exhaustivite_tab(df_presence: pd.DataFrame, config: ScoringConfig) -> None:
    st.header("📅 Contrôle d'Exhaustivité — Calendrier de Présence")

    if df_presence.empty:
        st.error("Données de présence indisponibles. Vérifier le fichier Pointage.")
        return

    builder = ExhaustiviteBuilder(config=config)

    # ── Filtres ───────────────────────────────────────────────────────
    col_f1, col_f2, col_f3, col_f4 = st.columns([2, 1, 2, 1])

    with col_f1:
        equipes_dispo = builder.get_equipes_list(df_presence)
        equipe_filter = st.multiselect(
            "🏢 Équipe(s)",
            options=equipes_dispo,
            default=equipes_dispo[:1] if equipes_dispo else [],
        )

    with col_f2:
        # ── NOUVEAU : sélection Année ─────────────────────────────────
        annees_dispo = _get_annees(df_presence)
        annee_sel = st.selectbox(
            "📅 Année",
            options=["Toutes"] + annees_dispo,
            index=len(annees_dispo) if annees_dispo else 0,
        )

    with col_f3:
        # Mois filtrés selon l'année sélectionnée
        mois_dispo = _get_mois_filtered(df_presence, annee_sel)
        mois_sel = st.selectbox(
            "📆 Mois",
            options=["Tous"] + mois_dispo,
            index=len(mois_dispo) if mois_dispo else 0,
        )

    with col_f4:
        vue_mode = st.radio("Vue", options=["Individuelle", "Équipe"], horizontal=True)

    # ── Filtrage ──────────────────────────────────────────────────────
    df_filtered = _apply_filters(
        df_presence, builder,
        equipes=equipe_filter if equipe_filter else None,
        annee=annee_sel,
        mois=mois_sel,
    )

    if df_filtered.empty:
        st.warning("Aucune donnée pour cette sélection.")
        return

    # ── Métriques ─────────────────────────────────────────────────────
    _render_metrics(df_filtered, builder)

    st.markdown("---")

    # ── Calendrier ────────────────────────────────────────────────────
    if mois_sel == "Tous":
        st.info("💡 Sélectionnez un mois spécifique pour le calendrier détaillé.")
        _render_monthly_summary_chart(df_filtered, builder.compute_daily_stats(df_filtered))
    else:
        try:
            use_nom = vue_mode == "Individuelle"
            pivot_heures, status_matrix = builder.build_pivot_calendar(df_filtered, use_nom=use_nom)
            _render_heatmap_calendar(pivot_heures, status_matrix)
        except Exception as exc:
            st.error(f"Erreur d'affichage : {exc}")

    st.markdown("---")
    _render_export_section(df_presence, builder)


# ══════════════════════════════════════════════════════════════════════
# HELPERS FILTRES
# ══════════════════════════════════════════════════════════════════════

def _get_annees(df: pd.DataFrame) -> list[str]:
    if "date" not in df.columns:
        return []
    dates = pd.to_datetime(df["date"], errors="coerce").dropna()
    return sorted(set(str(d.year) for d in dates), reverse=True)


def _get_mois_filtered(df: pd.DataFrame, annee: str) -> list[str]:
    if "date" not in df.columns:
        return []
    dates = pd.to_datetime(df["date"], errors="coerce").dropna()
    if annee != "Toutes":
        dates = dates[dates.dt.year == int(annee)]
    return sorted(set(d.strftime("%Y-%m") for d in dates))


def _apply_filters(
    df: pd.DataFrame,
    builder: ExhaustiviteBuilder,
    equipes: list[str] | None,
    annee: str,
    mois: str,
) -> pd.DataFrame:
    """Applique les filtres équipe + année + mois."""
    df = df.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")

    if equipes and "equipe_nom" in df.columns:
        df = df[df["equipe_nom"].isin(equipes)]

    if annee != "Toutes":
        df = df[df["date"].dt.year == int(annee)]

    if mois != "Tous":
        df = df[df["date"].dt.to_period("M").astype(str) == mois]

    return df


# ══════════════════════════════════════════════════════════════════════
# RENDER FONCTIONS (inchangées)
# ══════════════════════════════════════════════════════════════════════

def _render_metrics(df_filtered: pd.DataFrame, builder: ExhaustiviteBuilder) -> None:
    daily_stats = builder.compute_daily_stats(df_filtered)
    if not daily_stats.empty:
        ouvrable = daily_stats[~daily_stats["est_weekend"]]
        if not ouvrable.empty:
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Taux de présence", f"{ouvrable['taux_presence'].mean():.1%}")
            m2.metric("Absences total", int(ouvrable["nb_absents"].sum()))
            m3.metric("Jours > 8h", int(ouvrable["nb_excessifs"].sum()))
            m4.metric("Effectif", df_filtered["salarie_nom"].nunique())


def _render_heatmap_calendar(pivot_heures: pd.DataFrame, status_matrix: pd.DataFrame) -> None:
    """Heatmap inchangé — affiche les heures dans les cellules."""
    dates       = [str(d.date()) if hasattr(d, "date") else str(d) for d in pivot_heures.columns]
    techniciens = list(pivot_heures.index)

    map_val = {
        PresenceStatus.PRESENT:      1,
        PresenceStatus.ABSENT:       0,
        PresenceStatus.EXCESSIF:     2,
        PresenceStatus.WEEKEND:      -1,
        PresenceStatus.FERIE:        -2,
        PresenceStatus.NON_CONCERNE: -3,
    }

    z_colors, z_text, customdata = [], [], []
    for tech in techniciens:
        row_colors, row_text, row_custom = [], [], []
        for col_date in pivot_heures.columns:
            heures = pivot_heures.loc[tech, col_date]
            statut = status_matrix.loc[tech, col_date]
            row_colors.append(map_val.get(statut, -3))
            row_text.append(f"{heures:.1f}h" if heures > 0 else "")
            row_custom.append([heures, STATUS_LABELS.get(statut, statut)])
        z_colors.append(row_colors)
        z_text.append(row_text)
        customdata.append(row_custom)

    fig = go.Figure(data=go.Heatmap(
        z=z_colors, x=dates, y=techniciens,
        text=z_text,
        texttemplate="%{text}",
        textfont={"size": 9, "color": "black"},
        customdata=customdata,
        hovertemplate="<b>%{y}</b><br>%{x}<br>%{customdata[0]:.1f}h — %{customdata[1]}<extra></extra>",
        colorscale=[
            [0.0,   STATUS_COLORS[PresenceStatus.NON_CONCERNE]],
            [0.167, STATUS_COLORS[PresenceStatus.FERIE]],
            [0.333, STATUS_COLORS[PresenceStatus.WEEKEND]],
            [0.5,   STATUS_COLORS[PresenceStatus.ABSENT]],
            [0.667, STATUS_COLORS[PresenceStatus.PRESENT]],
            [1.0,   STATUS_COLORS[PresenceStatus.EXCESSIF]],
        ],
        showscale=False, xgap=1, ygap=1,
    ))
    fig.update_layout(
        title="Calendrier de présence — heures pointées par technicien",
        xaxis_title="Date", yaxis_title="Technicien",
        height=max(400, 30 * len(techniciens) + 150),
        margin=dict(l=200, r=20, t=60, b=60),
        plot_bgcolor="white",
        xaxis=dict(tickangle=-45, tickfont=dict(size=10)),
        yaxis=dict(tickfont=dict(size=10)),
    )
    st.plotly_chart(fig, use_container_width=True)

    col_leg = st.columns(len(STATUS_LABELS))
    for i, (status, label) in enumerate(STATUS_LABELS.items()):
        with col_leg[i]:
            color = STATUS_COLORS[status]
            st.markdown(
                f'<div style="background:{color};border-radius:4px;padding:4px 8px;'
                f'text-align:center;font-size:12px;color:{"white" if status == PresenceStatus.ABSENT else "black"}">'
                f'{label}</div>',
                unsafe_allow_html=True,
            )


def _render_monthly_summary_chart(df_filtered: pd.DataFrame, daily_stats: pd.DataFrame) -> None:
    if daily_stats.empty:
        return
    import plotly.express as px
    fig = px.bar(
        daily_stats[~daily_stats["est_weekend"]],
        x="date", y="taux_presence",
        color="taux_presence",
        color_continuous_scale=["#E74C3C", "#F39C12", "#2ECC71"],
        title="Taux de présence journalier (jours ouvrables)",
        labels={"taux_presence": "Taux présence", "date": "Date"},
    )
    fig.update_layout(yaxis_tickformat=".0%", height=350, showlegend=False)
    st.plotly_chart(fig, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════
# EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════

def _render_export_section(df_presence: pd.DataFrame, builder: ExhaustiviteBuilder) -> None:
    st.subheader("📥 Export Suivi de Présence")
    st.caption("Une feuille par équipe — P = présent, vide = absent.")

    col1, col2, col3 = st.columns([2, 1, 2])

    with col1:
        equipes = st.multiselect(
            "Équipes à exporter",
            options=builder.get_equipes_list(df_presence),
            default=builder.get_equipes_list(df_presence),
            key="exp_eq",
        )

    with col2:
        # ── NOUVEAU : sélection Année export ─────────────────────────
        annees_exp = ["Toutes"] + _get_annees(df_presence)
        annee_exp = st.selectbox("Année", annees_exp, key="exp_annee")

    with col3:
        # Trimestres + mois filtrés selon l'année
        mois_exp_dispo = _get_mois_filtered(df_presence, annee_exp)
        trimestres_exp = _get_trimestres_from_mois(mois_exp_dispo)
        periodes = ["Toute la période"] + trimestres_exp + mois_exp_dispo
        periode = st.selectbox("Période", periodes, key="exp_per")

    if not equipes:
        st.warning("Sélectionnez au moins une équipe.")
        return

    df_prev = _filter_for_export(df_presence, equipes, annee_exp, periode)
    nb_tech  = df_prev["salarie_nom"].nunique() if "salarie_nom" in df_prev.columns else 0
    nb_jours = df_prev["date"].nunique() if "date" in df_prev.columns else 0
    st.info(
        f"**Périmètre** : {len(equipes)} équipe(s) · "
        f"{nb_tech} technicien(s) · {nb_jours} jour(s) · {periode}"
        + (f" · {annee_exp}" if annee_exp != "Toutes" else "")
    )

    if st.button("⬇️ Générer l'Excel", type="primary", key="export_btn"):
        with st.spinner("Construction du fichier…"):
            try:
                data = _build_excel_engine(df_presence, equipes, annee_exp, periode)
                nom  = f"Presence_{annee_exp}_{periode.replace(' ', '_').replace('/', '-')}.xlsx"
                st.download_button(
                    label="📄 Télécharger",
                    data=data,
                    file_name=nom,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="export_dl",
                )
                st.success(f"✅ {len(equipes)} feuille(s) équipe + 1 feuille TOTAL générée(s)")
            except Exception as exc:
                st.error(f"Erreur : {exc}")


def _build_excel_engine(
    df_presence: pd.DataFrame,
    equipes: list[str],
    annee: str,
    periode: str,
) -> bytes:
    """
    Génère l'Excel multi-feuilles.
    - 1 feuille par équipe
    - 1 feuille TOTAL (agrégat de toutes les équipes sélectionnées)

    IMPORTANT : le pivot de présence est BINAIRE (1 = au moins 1 ligne de pointage,
    0 = absent) pour éviter les sommes anormales dues aux doublons de lignes.
    Les heures restent affichées dans le heatmap Streamlit, pas ici.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = _filter_for_export(df_presence, equipes, annee, periode)
    if df.empty:
        raise ValueError("Aucune donnée pour cette sélection.")

    # Couleurs
    NAVY  = "002060"
    BLEU  = "DDEBF7"
    VERT  = "C6EFCE"
    ROUGE = "FFC7CE"
    GRIS  = "D9D9D9"
    JAUNE = "FFF2CC"
    ORANGE = "FFC000"
    TOTAL_HDR = "1F3864"  # bleu foncé pour feuille TOTAL

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    output = BytesIO()

    # Collecte des données résumé pour la feuille TOTAL
    total_rows: list[dict] = []

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # ── 1. Feuilles par équipe ────────────────────────────────────
        for equipe in sorted(equipes):
            df_eq = df[df["equipe_nom"] == equipe].copy()
            if df_eq.empty:
                continue

            df_eq["date"] = pd.to_datetime(df_eq["date"]).dt.normalize()
            tous_jours    = pd.date_range(df_eq["date"].min(), df_eq["date"].max(), freq="D")

            # ── Pivot BINAIRE ─────────────────────────────────────────
            # On détecte la présence par l'existence d'au moins une ligne
            # de pointage, indépendamment des heures (évite les doublons).
            pivot_presence = _build_binary_pivot(df_eq, tous_jours)

            techniciens     = list(pivot_presence.index)
            nb_jours_ouvres = _nb_jours_ouvres(tous_jours[0], tous_jours[-1])
            feries          = _get_feries(tous_jours[0], tous_jours[-1])

            ws = writer.book.create_sheet(title=equipe[:31])

            _write_sheet(
                ws=ws,
                titre=f"SUIVI PRÉSENCE — {equipe} ({periode})",
                techniciens=techniciens,
                tous_jours=tous_jours,
                pivot_presence=pivot_presence,
                nb_jours_ouvres=nb_jours_ouvres,
                feries=feries,
                hdr_color=NAVY,
                border=border,
                BLEU=BLEU, VERT=VERT, ROUGE=ROUGE, GRIS=GRIS,
                JAUNE=JAUNE, ORANGE=ORANGE,
            )

            # Prépare les lignes pour TOTAL
            for tech in techniciens:
                presents = int(pivot_presence.loc[tech].sum())
                taux = round(presents / nb_jours_ouvres, 4) if nb_jours_ouvres > 0 else 0
                total_rows.append({
                    "Équipe": equipe,
                    "Technicien": tech,
                    "Jours présents": presents,
                    "Jours ouvrés": nb_jours_ouvres,
                    "Taux présence": taux,
                })

        # ── 2. Feuille TOTAL ──────────────────────────────────────────
        _write_total_sheet(
            writer=writer,
            total_rows=total_rows,
            periode=periode,
            annee=annee,
            hdr_color=TOTAL_HDR,
            border=border,
            VERT=VERT, ROUGE=ROUGE, JAUNE=JAUNE, ORANGE=ORANGE,
        )

    return output.getvalue()


def _build_binary_pivot(df_eq: pd.DataFrame, tous_jours: pd.DatetimeIndex) -> pd.DataFrame:
    """
    Pivot binaire technicien × jour.
    Valeur = 1 si au moins une ligne de pointage existe ce jour, 0 sinon.
    Insensible aux doublons et aux valeurs d'heures anormales.
    """
    # Marquer chaque ligne comme "présence détectée = 1"
    df_eq = df_eq.copy()
    df_eq["_present"] = 1

    pivot = df_eq.pivot_table(
        index="salarie_nom",
        columns="date",
        values="_present",
        aggfunc="max",          # max(1, 1, 1...) = 1 — insensible aux doublons
    ).reindex(columns=tous_jours, fill_value=0).fillna(0).astype(int)

    return pivot


def _write_sheet(
    ws,
    titre: str,
    techniciens: list[str],
    tous_jours: pd.DatetimeIndex,
    pivot_presence: pd.DataFrame,
    nb_jours_ouvres: int,
    feries: set,
    hdr_color: str,
    border,
    BLEU, VERT, ROUGE, GRIS, JAUNE, ORANGE,
) -> None:
    """Écrit une feuille équipe dans le workbook."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    # Titre
    ws["A1"] = titre
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill      = PatternFill("solid", fgColor=hdr_color)
    ws["A1"].alignment = Alignment(horizontal="left")

    # En-tête colonne technicien
    cell = ws.cell(row=3, column=1)
    cell.value     = "Technicien"
    cell.font      = Font(bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", fgColor=hdr_color)
    cell.alignment = Alignment(horizontal="left")
    cell.border    = border
    ws.column_dimensions["A"].width = 30

    # En-têtes dates
    for c_idx, jour in enumerate(tous_jours, start=2):
        cell           = ws.cell(row=3, column=c_idx)
        cell.value     = f"{jour.strftime('%d/%m')}\n{_JOURS_FR[jour.weekday()]}"
        cell.fill      = PatternFill("solid", fgColor=BLEU)
        cell.font      = Font(bold=True, size=8)
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(c_idx)].width = 7

    # En-têtes colonnes résumé
    res_start = len(tous_jours) + 2
    for i, label in enumerate(["Jours présents", "Jours ouvrés", "Taux présence"]):
        col_idx        = res_start + i
        cell           = ws.cell(row=3, column=col_idx)
        cell.value     = label
        cell.fill      = PatternFill("solid", fgColor=ORANGE)
        cell.font      = Font(bold=True, size=9, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # Données
    for r_idx, tech in enumerate(techniciens, start=4):
        cell_tech            = ws.cell(row=r_idx, column=1)
        cell_tech.value      = tech
        cell_tech.font       = Font(bold=True, size=9)
        cell_tech.fill       = PatternFill("solid", fgColor="EEF3FA")
        cell_tech.alignment  = Alignment(horizontal="left", vertical="center")
        cell_tech.border     = border
        ws.row_dimensions[r_idx].height = 18

        presents = 0

        for c_idx, jour in enumerate(tous_jours, start=2):
            cell           = ws.cell(row=r_idx, column=c_idx)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = border

            # Lecture directe du pivot binaire
            val        = int(pivot_presence.loc[tech, jour])
            is_weekend = jour.weekday() >= 5
            is_ferie   = jour.date() in feries

            if is_weekend:
                cell.fill = PatternFill("solid", fgColor=GRIS)
            elif is_ferie:
                # Jour férié : affiché en bleu clair, non compté dans le dénominateur
                cell.value = "F"
                cell.fill  = PatternFill("solid", fgColor="BDD7EE")
                cell.font  = Font(italic=True, size=8, color="1F497D")
            elif val == 1:
                cell.value = "P"
                cell.fill  = PatternFill("solid", fgColor=VERT)
                cell.font  = Font(bold=True, size=9, color="276221")
                presents  += 1
            else:
                cell.fill = PatternFill("solid", fgColor=ROUGE)

        # Colonnes résumé
        taux = round(presents / nb_jours_ouvres, 4) if nb_jours_ouvres > 0 else 0
        for i, (val, label) in enumerate([
            (presents,        "Jours présents"),
            (nb_jours_ouvres, "Jours ouvrés"),
            (taux,            "Taux présence"),
        ]):
            col_idx        = res_start + i
            cell           = ws.cell(row=r_idx, column=col_idx)
            cell.value     = val
            cell.alignment = Alignment(horizontal="center")
            cell.border    = border
            if label == "Taux présence":
                cell.number_format = "0%"
                if taux >= 0.9:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(bold=True, color="276221", size=9)
                elif taux >= 0.7:
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                    cell.font = Font(bold=True, color="9C6500", size=9)
                else:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(bold=True, color="9C0006", size=9)
            else:
                cell.fill = PatternFill("solid", fgColor=JAUNE)

    ws.row_dimensions[3].height = 36
    ws.freeze_panes = "B4"


def _write_total_sheet(
    writer,
    total_rows: list[dict],
    periode: str,
    annee: str,
    hdr_color: str,
    border,
    VERT, ROUGE, JAUNE, ORANGE,
) -> None:
    """
    Feuille TOTAL : résumé agrégé de toutes les équipes sélectionnées.
    Bloc par équipe + ligne TOTAL GLOBAL en bas.
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    if not total_rows:
        return

    ws = writer.book.create_sheet(title="TOTAL", index=0)  # en premier onglet

    label_periode = f"{periode}" + (f" · {annee}" if annee != "Toutes" else "")

    # Titre
    ws["A1"] = f"EXHAUSTIVITÉ GLOBALE — {label_periode}"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill      = PatternFill("solid", fgColor=hdr_color)
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.merge_cells("A1:E1")

    # En-têtes tableau
    headers = ["Équipe", "Technicien", "Jours présents", "Jours ouvrés", "Taux présence"]
    col_widths = [25, 30, 14, 12, 14]
    for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell           = ws.cell(row=3, column=c)
        cell.value     = h
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=hdr_color)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(c)].width = w

    # Données par équipe (triées)
    df_total = pd.DataFrame(total_rows).sort_values(["Équipe", "Technicien"])
    current_equipe = None
    r_idx = 4

    for _, row in df_total.iterrows():
        # Séparateur de groupe équipe
        if row["Équipe"] != current_equipe:
            current_equipe = row["Équipe"]
            # Ligne sous-titre équipe
            cell_eq           = ws.cell(row=r_idx, column=1)
            cell_eq.value     = f"▶ {current_equipe}"
            cell_eq.font      = Font(bold=True, color="FFFFFF", size=10)
            cell_eq.fill      = PatternFill("solid", fgColor="2E75B6")
            cell_eq.alignment = Alignment(horizontal="left")
            ws.merge_cells(f"A{r_idx}:E{r_idx}")
            ws.row_dimensions[r_idx].height = 20
            r_idx += 1

        # Ligne technicien
        taux = row["Taux présence"]
        vals = [row["Équipe"], row["Technicien"], row["Jours présents"],
                row["Jours ouvrés"], taux]
        for c, val in enumerate(vals, start=1):
            cell           = ws.cell(row=r_idx, column=c)
            cell.value     = val
            cell.alignment = Alignment(horizontal="center" if c > 2 else "left")
            cell.border    = border
            cell.font      = Font(size=9)
            if c == 5:  # Taux présence
                cell.number_format = "0%"
                if taux >= 0.9:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(bold=True, color="276221", size=9)
                elif taux >= 0.7:
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                    cell.font = Font(bold=True, color="9C6500", size=9)
                else:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(bold=True, color="9C0006", size=9)
            else:
                cell.fill = PatternFill("solid", fgColor="F5F5F5" if r_idx % 2 == 0 else "FFFFFF")
        ws.row_dimensions[r_idx].height = 16
        r_idx += 1

    # ── Ligne TOTAL GLOBAL ────────────────────────────────────────────
    r_idx += 1  # ligne vide de séparation

    total_presents = df_total["Jours présents"].sum()
    # Jours ouvrés = max par technicien (ils ont tous la même valeur dans la période)
    total_ouvres   = df_total["Jours ouvrés"].sum()  # somme car c'est par technicien
    taux_global    = round(total_presents / total_ouvres, 4) if total_ouvres > 0 else 0

    labels_total = ["TOUTES ÉQUIPES", f"{df_total['Technicien'].nunique()} techniciens",
                    total_presents, total_ouvres, taux_global]
    for c, val in enumerate(labels_total, start=1):
        cell           = ws.cell(row=r_idx, column=c)
        cell.value     = val
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center" if c > 2 else "left")
        cell.border    = border
        if c == 5:
            cell.number_format = "0%"
    ws.row_dimensions[r_idx].height = 22

    ws.freeze_panes = "A4"


# ══════════════════════════════════════════════════════════════════════
# HELPERS EXPORT
# ══════════════════════════════════════════════════════════════════════

def _get_trimestres(df: pd.DataFrame) -> list[str]:
    if "date" not in df.columns:
        return []
    dates = pd.to_datetime(df["date"], errors="coerce").dropna()
    return sorted(set(f"T{(d.month - 1) // 3 + 1} {d.year}" for d in dates))


def _get_trimestres_from_mois(mois_list: list[str]) -> list[str]:
    """Déduit les trimestres disponibles depuis une liste de 'YYYY-MM'."""
    trimestres = set()
    for m in mois_list:
        try:
            y, mo = int(m[:4]), int(m[5:7])
            trimestres.add(f"T{(mo - 1) // 3 + 1} {y}")
        except ValueError:
            pass
    return sorted(trimestres)


def _filter_for_export(
    df_presence: pd.DataFrame,
    equipes: list[str],
    annee: str,
    periode: str,
) -> pd.DataFrame:
    df = df_presence.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")

    if equipes and "equipe_nom" in df.columns:
        df = df[df["equipe_nom"].isin(equipes)]

    # Filtre année
    if annee != "Toutes":
        df = df[df["date"].dt.year == int(annee)]

    # Filtre période (trimestre ou mois)
    if periode != "Toute la période":
        if periode.startswith("T"):
            q = int(periode[1])
            y = int(periode.split()[1])
            df = df[
                (df["date"].dt.year == y) &
                (((df["date"].dt.month - 1) // 3 + 1) == q)
            ]
        else:
            df = df[df["date"].dt.to_period("M").astype(str) == periode]

    return df
